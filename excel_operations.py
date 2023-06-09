import win32com.client
import openpyxl as xl
from openpyxl.styles import Alignment, PatternFill, Side, Border
import os
import xlsxwriter


def add_rows(work_book):
    sheet = work_book.active
    # Add new rows and set correct values
    rowIndex = 2
    rowMax = sheet.max_row + 1
    while rowIndex < rowMax:
        cell = sheet.cell(rowIndex, 6).value
        if cell == 28:
            # Add new row
            sheet.insert_rows(rowIndex + 1)
            sheet.cell(rowIndex, 6).value = 10
            # Operations on operation text.1 for current row
            operation_text = str(sheet.cell(rowIndex, 5).value)
            operation_text = operation_text[:3] + "10" + operation_text[5:]
            sheet.cell(rowIndex, 2).value = sheet.cell(rowIndex, 2).value + 30
            sheet.cell(rowIndex, 3).value = sheet.cell(rowIndex, 3).value + 30
            sheet.cell(rowIndex, 5).value = operation_text
            sheet.cell(rowIndex, 16).value = operation_text
            # create Tekst oper.10
            operation_text = f'po sklejeniu na 28mm sciąć na {sheet.cell(rowIndex, 2).value} x {sheet.cell(rowIndex, 3).value}'
            sheet.cell(rowIndex, 25).value = operation_text
            # Copy values from row above
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row=rowIndex + 1, column=column).value = sheet.cell(row=rowIndex, column=column).value
            sheet.cell(rowIndex + 1, 6).value = 18
            # Operations on operation text.1 for new row
            operation_text = str(sheet.cell(rowIndex + 1, 5).value)
            operation_text = operation_text[:3] + "18" + operation_text[5:]
            sheet.cell(rowIndex + 1, 5).value = operation_text
            sheet.cell(rowIndex + 1, 16).value = operation_text
            #
            rowIndex += 1
            rowMax += 1
        if cell == 36:
            # Add new row
            sheet.insert_rows(rowIndex + 1)
            sheet.cell(rowIndex, 6).value = 18
            # Operations on operation text.1 for current row
            operation_text = str(sheet.cell(rowIndex, 16).value)
            operation_text = operation_text[:3] + "18" + operation_text[5:]
            sheet.cell(rowIndex, 2).value = sheet.cell(rowIndex, 2).value + 30
            sheet.cell(rowIndex, 3).value = sheet.cell(rowIndex, 3).value + 30
            sheet.cell(rowIndex, 5).value = operation_text
            sheet.cell(rowIndex, 16).value = operation_text
            # create Tekst oper.10
            operation_text = f'po sklejeniu na 36mm sciąć na {sheet.cell(rowIndex, 2).value} x {sheet.cell(rowIndex, 3).value}'
            sheet.cell(rowIndex, 25).value = operation_text
            # Copy values from row above
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row=rowIndex + 1, column=column).value = sheet.cell(row=rowIndex, column=column).value
            #
            rowIndex += 1
            rowMax += 1
        rowIndex += 1  # Move to the newly inserted row


def save_excel_to_pdf(wb_file_name, original_file_name):
    # get file_path
    file_path = os.path.join(os.getcwd(), wb_file_name)

    # Convert Excel file to PDF using win32com
    excel = win32com.client.Dispatch('Excel.Application')
    excel.DisplayAlerts = False  # Disable the prompt to save changes

    workbook = excel.Workbooks.Open(file_path)

    # Set properties for printing without margins
    active_sheet = workbook.ActiveSheet
    active_sheet.PageSetup.FitToPagesWide = 1
    active_sheet.PageSetup.FitToPagesTall = False
    active_sheet.PageSetup.LeftMargin = 0
    active_sheet.PageSetup.RightMargin = 0
    active_sheet.PageSetup.TopMargin = 0
    active_sheet.PageSetup.BottomMargin = 0

    # Save to pdf
    directory_path = os.path.dirname(file_path)
    pdf_file_path = directory_path + '\\' + original_file_name + '.pdf'
    workbook.ExportAsFixedFormat(0, pdf_file_path, 1, 0, 0)

    # Close the Excel file
    workbook.Close()
    excel.Quit()


import openpyxl

from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def sub_collumns(file_path):
    # Wczytanie pliku Excel
    # workbook = load_workbook(file_path)
    workbook = file_path

    # Wybór odpowiedniego arkusza
    worksheet = workbook.active

    # Zapis nagłówków
    header1 = worksheet.cell(row=2, column=18).value
    header2 = worksheet.cell(row=2, column=21).value
    header3 = worksheet.cell(row=2, column=22).value

    worksheet.insert_rows(1)

    # Usunięcie kolumn niepotrzebnych do formatowania
    columns_to_delete = []
    for column in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column)
        if column_letter not in ['A', 'B', 'C', 'E', 'F', 'Q', 'X', "Y"]:
            columns_to_delete.append(column)

    for column_index in sorted(columns_to_delete, reverse=True):
        worksheet.delete_cols(column_index, 1)

    # Ustawienie stylu dla pozostałych kolumn
    table_style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                                          showLastColumn=False, showRowStripes=True,
                                                          showColumnStripes=True)
    table = openpyxl.worksheet.table.Table(ref=f"A2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}",
                                           displayName="MyTable", tableStyleInfo=table_style)
    worksheet.add_table(table)

    # Utworzenie stylu dla nagłówków
    header_style = NamedStyle(name="header_style")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.fill = PatternFill(start_color=Color(rgb="EBF1DE"), end_color=Color(rgb="EBF1DE"), fill_type="solid")
    header_style.border = Border(left=Side(style="thin", color=Color(rgb="9BBB59")),
                                 right=Side(style="thin", color=Color(rgb="9BBB59")),
                                 top=Side(style="thin", color=Color(rgb="9BBB59")),
                                 bottom=Side(style="thin", color=Color(rgb="9BBB59")))

    # Wstawienie nagłówków do pierwszego wiersza tabeli
    worksheet.cell(row=1, column=1, value=header1).style = header_style
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    worksheet.cell(row=1, column=4, value=header2).style = header_style
    worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    worksheet.cell(row=1, column=6, value=header3).style = header_style
    worksheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)

    # Szerokość kolumn
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 8
    worksheet.column_dimensions['C'].width = 8
    worksheet.column_dimensions['D'].width = 17.1
    worksheet.column_dimensions['E'].width = 7
    worksheet.column_dimensions['F'].width = 13
    worksheet.column_dimensions['G'].width = 13
    worksheet.column_dimensions['H'].width = 37

    # row_width = worksheet.column_dimensions['A'].width
    # row_width += worksheet.column_dimensions['B'].width
    # row_width += worksheet.column_dimensions['C'].width
    # row_width += worksheet.column_dimensions['D'].width
    # row_width += worksheet.column_dimensions['E'].width
    # row_width += worksheet.column_dimensions['F'].width
    # row_width += worksheet.column_dimensions['G'].width
    # row_width += worksheet.column_dimensions['H'].width
    # print(row_width)

    # Ustawienie stylu dla wierszy
    row_style1 = NamedStyle(name="row_style1")
    row_style1.fill = PatternFill(start_color=Color(rgb="EBF1DE"), end_color=Color(rgb="EBF1DE"), fill_type="solid")
    row_style1.border = Border(left=Side(style="thin", color=Color(rgb="9BBB59")),
                               right=Side(style="thin", color=Color(rgb="9BBB59")),
                               top=Side(style="thin", color=Color(rgb="9BBB59")),
                               bottom=Side(style="thin", color=Color(rgb="9BBB59")))

    row_style2 = NamedStyle(name="row_style2")
    row_style2.fill = PatternFill(start_color=Color(rgb="FFFFFF"), end_color=Color(rgb="FFFFFF"), fill_type="solid")
    row_style2.border = Border(left=Side(style="thin", color=Color(rgb="9BBB59")),
                               right=Side(style="thin", color=Color(rgb="9BBB59")),
                               top=Side(style="thin", color=Color(rgb="9BBB59")),
                               bottom=Side(style="thin", color=Color(rgb="9BBB59")))

    for row in worksheet.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.style = row_style1
        else:
            for cell in row:
                cell.style = row_style2

    # Zapisanie zmian do pliku
    workbook.save('Book1__________________Operations.xlsx')


def delete_file(file_name):
    os.remove(file_name)
